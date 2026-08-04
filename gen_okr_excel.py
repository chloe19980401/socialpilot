# -*- coding: utf-8 -*-
"""
生成《市场部 8月 OKR》Excel 表（每人 3 个目标O，每个O 4 个关键结果KR，SMART）。
用法：  python gen_okr_excel.py
自动安装 openpyxl（若未装），在脚本同目录生成： 市场部8月OKR.xlsx
"""
import os, sys, subprocess
try:
    import openpyxl
except ImportError:
    print("未检测到 openpyxl，正在自动安装…")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------------- OKR 内容 ----------------
# 每个 kr = (关键结果, 衡量/目标值, 截止)
DATA = [
 {"person":"李铧燕","role":"自建站 + 视频/设计 + 社媒统筹/系统","objectives":[
   {"o":"O1 自建站(wonlyglobal.com)运营","krs":[
     ("Google 收录页数 30 → ≥ 40","数量","8/31"),
     ("Google 月展现 50 → ≥ 300","数量","8/31"),
     ("非品牌词 ≥ 5 个进前 20 名","排名","8/31"),
     ("SEO 来源有效询盘 0 → ≥ 3；全部页面上线 + 俄语多语言试点","数量/交付","8/31")]},
   {"o":"O2 安装视频本地化","krs":[
     ("西班牙代理-木门&防盗门安装视频 ≥ 2 条交付","数量","8/20"),
     ("全程 AI 翻译配音，单条本地化 API 成本 ≈ 0","成本","8/31"),
     ("双语字幕 + 画面原文替换覆盖 100%","覆盖率","8/31"),
     ("代理确认一次通过率 ≥ 90%","比率","8/31")]},
   {"o":"O3 卖点视频与展会物料","krs":[
     ("业务部门卖点视频 ≥ 19 条完成","数量","8/31"),
     ("竖屏 1080×1920、单条 ≤ 60s，每周 ≥ 5 条","规格/数量","8/31"),
     ("迪拜展会设计（与赵琳）+ 展会宣传视频 ≥ 1 条","交付","8/31"),
     ("交付社媒复用 ≥ 10 条、素材 100% 归档","数量/覆盖","8/31")]},
   {"o":"O4 社媒中台运行与流程线上化","krs":[
     ("四平台数据自动同步日更、成功率 ≥ 95%","比率","8/31"),
     ("飞书日报 + 上传提醒 100% 正常推送","比率","8/31"),
     ("排期→审批→发布→复盘 100% 系统内完成、发帖自动回填“已发布”","覆盖率","8/31"),
     ("设计台承接 100% 设计需求、绩效看板自动统计可导出","覆盖率","8/31")]},
   {"o":"O5 数据驱动与 AI 赋能","krs":[
     ("沉淀 ≥ 45 竞品 + ≥ 300 条指标快照、月度数据简报 ≥ 1 份","数量","8/31"),
     ("支撑社媒组达成发帖/粉丝/曝光/询盘四项 KPI","达成","8/31"),
     ("AI 工具市场部覆盖 100%（6 人）","覆盖率","8/31"),
     ("沉淀 SOP/使用手册 ≥ 1 份、系统可用性 ≥ 99%","交付/比率","8/31")]}]},

 {"person":"周宗莉","role":"社媒运营·内容产出","objectives":[
   {"o":"O1 内容稳定高产","krs":[
     ("各平台发帖 ≥ 40 条/月（YT/IG/TikTok/FB 各 ≥ 40）","数量","8/31"),
     ("每周产出重点原创内容 ≥ 2 条，8月累计 ≥ 8 条","数量","每周"),
     ("内容排期按时发布率 ≥ 95%、逾期 0 条","比率","8/31"),
     ("帖子 100% 录入内容中心并关联排期","覆盖率","8/31")]},
   {"o":"O2 内容质量与选题","krs":[
     ("建立海外内容选题库 ≥ 30 条并滚动更新","数量","8/15"),
     ("爆款内容（单条曝光 ≥ 1,000）≥ 3 条","数量","8/31"),
     ("内容平均互动率 ≥ 3%","比率","8/31"),
     ("每周 1 次内容复盘，沉淀可复用模板 ≥ 4 个","数量","每周")]},
   {"o":"O3 协同与规范","krs":[
     ("对设计/视频需求 100% 提前 3 天提报","及时率","8/31"),
     ("产出发帖 SOP / 文案规范 ≥ 1 份","交付","8/20"),
     ("跨平台内容复用率 ≥ 50%","比率","8/31"),
     ("内容素材 100% 归档","覆盖率","8/31")]}]},

 {"person":"周雨晴","role":"社媒运营·粉丝增长/社群","objectives":[
   {"o":"O1 粉丝增长","krs":[
     ("全平台总粉丝 10 → ≥ 100","数量","8/31"),
     ("各平台净增粉丝 ≥ 20","数量","8/31"),
     ("粉丝月流失率 ≤ 5%","比率","8/31"),
     ("关注引导/互动获粉动作 ≥ 30 次/周","频次","每周")]},
   {"o":"O2 曝光与触达","krs":[
     ("全平台月曝光 420 → ≥ 30,000","数量","8/31"),
     ("平均单条曝光 ≥ 500","数量","8/31"),
     ("主页/落地链接点击 ≥ 50 次","数量","8/31"),
     ("蹭热点/话题标签内容 ≥ 8 条","数量","8/31")]},
   {"o":"O3 社群互动","krs":[
     ("评论/私信 24h 内回复率 ≥ 90%","比率","8/31"),
     ("主动互动（点赞/评论潜客同行）≥ 100 次/周","频次","每周"),
     ("发起互动活动/话题 ≥ 2 次","数量","8/31"),
     ("建立粉丝画像/互动记录表 ≥ 1 份","交付","8/20")]}]},

 {"person":"徐一诺","role":"社媒运营·数据/竞品/转化","objectives":[
   {"o":"O1 转化与获客","krs":[
     ("社媒来源有效询盘 ≥ 3 条","数量","8/31"),
     ("落地页/表单转化线索 ≥ 5 条","数量","8/31"),
     ("询盘 24h 内响应率 100%","比率","8/31"),
     ("建立社媒→询盘跟进表，闭环记录 100%","覆盖率","8/31")]},
   {"o":"O2 竞品与情报","krs":[
     ("持续监测竞品 ≥ 45 个","数量","8/31"),
     ("输出竞品对标简报 ≥ 1 份/月","交付","8/25"),
     ("提炼可借鉴打法 ≥ 5 条并落地 ≥ 2 条","数量","8/31"),
     ("KOL/红人储备 ≥ 10 家、发起洽谈 ≥ 3 家","数量","8/31")]},
   {"o":"O3 数据分析","krs":[
     ("每周输出社媒数据周报 ≥ 4 期","数量","每周"),
     ("搭建平台数据看板，指标日更","交付","8/15"),
     ("平均互动率提升至 ≥ 3%","比率","8/31"),
     ("基于数据给出优化建议 ≥ 4 条/月并验证","数量","8/31")]}]},

 {"person":"赵琳","role":"设计","objectives":[
   {"o":"O1 海外画册物料交付","krs":[
     ("智能锁画册 36P 定稿","交付","8/15"),
     ("王力静音木门画册 56P 定稿","交付","8/25"),
     ("王力智能门窗画册 35P 定稿","交付","8/25"),
     ("SI 手册海外版（含审核）定稿","交付","8/31")]},
   {"o":"O2 展会与社媒设计支持","krs":[
     ("迪拜展会视觉设计（与李铧燕）完成","交付","8/31"),
     ("为社媒组供配图/海报 ≥ 20 张，交付 ≤ 2 个工作日","数量","8/31"),
     ("设计需求按时交付率 ≥ 90%、逾期 0","比率","8/31"),
     ("建立海外物料设计规范/模板库 ≥ 1 套","交付","8/20")]},
   {"o":"O3 效率与协同","krs":[
     ("全部设计任务在设计台线上流转、交付图 100% 归档","覆盖率","8/31"),
     ("单画册平均产出周期 ≤ 7 天","周期","8/31"),
     ("模板复用使重复类物料工时下降 ≥ 20%","比率","8/31"),
     ("需求 100% 有记录、单项返工 ≤ 1 次","比率","8/31")]}]},
]

# ---------------- 样式 ----------------
GOLD = "B8862A"; LIGHT = "F3EFE6"; GRAY = "D9D9D9"
thin = Side(style="thin", color="C8C8C8")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
FONT = "微软雅黑"
wrap_top = Alignment(wrap_text=True, vertical="top")
mid = Alignment(wrap_text=True, vertical="center", horizontal="center")

def build_sheet(ws, person):
    # 标题行：人员 + 岗位
    ws.merge_cells("A1:D1")
    t = ws.cell(1, 1, "%s  ·  %s   |   2026年8月 OKR" % (person["person"], person["role"]))
    t.fill = PatternFill("solid", fgColor=GOLD)
    t.font = Font(name=FONT, bold=True, color="FFFFFF", size=12)
    t.alignment = mid
    ws.row_dimensions[1].height = 28
    # 表头
    for c, h in enumerate(["目标 (O)", "关键结果 (KR)", "衡量/目标值", "截止"], 1):
        cell = ws.cell(2, c, h)
        cell.fill = PatternFill("solid", fgColor=GRAY)
        cell.font = Font(name=FONT, bold=True, size=10.5)
        cell.alignment = mid; cell.border = border
    r = 3
    for obj in person["objectives"]:
        o_start = r
        for kr, measure, due in obj["krs"]:
            ws.cell(r, 2, kr).alignment = wrap_top
            ws.cell(r, 3, measure).alignment = mid
            ws.cell(r, 4, due).alignment = mid
            for c in range(1, 5):
                ws.cell(r, c).border = border
                ws.cell(r, c).font = Font(name=FONT, size=10.5)
            ws.row_dimensions[r].height = 30
            r += 1
        ws.merge_cells(start_row=o_start, start_column=1, end_row=r - 1, end_column=1)
        oc = ws.cell(o_start, 1, obj["o"])
        oc.alignment = mid; oc.font = Font(name=FONT, bold=True, size=10.5, color="9C272E")
    for c, w in enumerate([24, 58, 14, 9], 1):
        ws.column_dimensions[chr(64 + c)].width = w
    ws.freeze_panes = "A3"


wb = Workbook()
for i, person in enumerate(DATA):
    ws = wb.active if i == 0 else wb.create_sheet()
    ws.title = person["person"][:31]  # 工作表名即人名（≤31字符）
    build_sheet(ws, person)

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "市场部8月OKR.xlsx")
try:
    wb.save(out)
except PermissionError:
    import time
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "市场部8月OKR_%s.xlsx" % time.strftime("%H%M%S"))
    wb.save(out)
    print("（原文件正在 Excel 里打开、被占用，已自动另存为新文件）")
print("已生成：", out)
